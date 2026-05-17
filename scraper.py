import requests
import re
import json
import csv
import os
import hashlib
from datetime import datetime
from bs4 import BeautifulSoup
import openpyxl
from urllib.parse import urljoin, urlparse

# ─── Duplicate Detection ──────────────────────────────────

seen_hashes = set()

def is_duplicate(text):
    h = hashlib.md5(text.encode()).hexdigest()
    if h in seen_hashes:
        return True
    seen_hashes.add(h)
    return False

def reset_duplicates():
    seen_hashes.clear()

# ─── Fetch Page ───────────────────────────────────────────

def fetch_page(url):
    print("[*] Fetching: " + url)
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"
        }
        response = requests.get(url, headers=headers, timeout=10)
        if response.status_code == 200:
            return response.text
        else:
            print("[-] Failed. Status code: " + str(response.status_code))
            return None
    except requests.exceptions.ConnectionError:
        print("[-] Could not connect to: " + url)
        return None
    except requests.exceptions.Timeout:
        print("[-] Connection timed out.")
        return None

# ─── Pagination ───────────────────────────────────────────

def find_next_page(soup, current_url):
    next_patterns = ['next', 'next page', '>', '»', 'load more']
    for a in soup.find_all('a', href=True):
        text = a.get_text().strip().lower()
        if any(pattern in text for pattern in next_patterns):
            next_url = urljoin(current_url, a['href'])
            return next_url
    for a in soup.find_all('a', href=True):
        classes = ' '.join(a.get('class', [])).lower()
        if 'next' in classes:
            return urljoin(current_url, a['href'])
    return None

def get_all_pages(start_url, max_pages=10):
    pages = []
    current_url = start_url
    page_num = 1
    print("\n[*] Starting multi-page scrape...")
    while current_url and page_num <= max_pages:
        print("[*] Scraping page " + str(page_num) + ": " + current_url)
        html = fetch_page(current_url)
        if not html:
            break
        pages.append((current_url, html))
        soup = BeautifulSoup(html, 'html.parser')
        next_url = find_next_page(soup, current_url)
        if next_url == current_url:
            break
        current_url = next_url
        page_num = page_num + 1
    print("[+] Scraped " + str(len(pages)) + " pages total.")
    return pages

# ─── Image Downloader ─────────────────────────────────────

def download_images(images, domain, max_images=50):
    folder = "images_" + domain.replace(".", "_")
    if not os.path.exists(folder):
        os.makedirs(folder)
    print("\n[*] Downloading images to: " + folder)
    downloaded = 0
    for img in images[:max_images]:
        url = img['url']
        try:
            response = requests.get(url, timeout=10)
            if response.status_code == 200:
                ext = url.split('.')[-1].split('?')[0]
                if ext.lower() not in ['jpg','jpeg','png','gif','webp','svg']:
                    ext = 'jpg'
                filename = os.path.join(folder, str(downloaded + 1) + "." + ext)
                with open(filename, 'wb') as f:
                    f.write(response.content)
                downloaded = downloaded + 1
                print("[+] Downloaded: " + filename)
        except:
            pass
    print("[+] Downloaded " + str(downloaded) + " images to: " + folder)
    return folder

# ─── Extractors ───────────────────────────────────────────

def extract_headlines(soup):
    headlines = []
    for tag in ['h1', 'h2', 'h3', 'h4']:
        for item in soup.find_all(tag):
            text = item.get_text().strip()
            if text and len(text) > 3 and not is_duplicate(text):
                headlines.append({
                    "type": tag.upper(),
                    "text": text
                })
    return headlines

def extract_links(soup, base_url):
    links = []
    for item in soup.find_all('a', href=True):
        href = item['href'].strip()
        text = item.get_text().strip()
        if href.startswith('http'):
            full_url = href
        elif href.startswith('/'):
            full_url = base_url.rstrip('/') + href
        elif href.startswith('#') or href.startswith('mailto'):
            continue
        else:
            full_url = base_url.rstrip('/') + '/' + href
        if not is_duplicate(full_url):
            links.append({
                "text": text if text else "No text",
                "url": full_url
            })
    return links

def extract_images(soup, base_url):
    images = []
    for item in soup.find_all('img'):
        src = item.get('src', '').strip()
        alt = item.get('alt', '').strip()
        if not src:
            continue
        if src.startswith('http'):
            full_url = src
        elif src.startswith('/'):
            full_url = base_url.rstrip('/') + src
        else:
            full_url = base_url.rstrip('/') + '/' + src
        if not is_duplicate(full_url):
            images.append({
                "alt": alt if alt else "No description",
                "url": full_url
            })
    return images

def extract_emails(html):
    email_pattern = re.compile(
        r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}')
    found = list(set(email_pattern.findall(html)))
    return [e for e in found if not is_duplicate(e)]

def extract_phone_numbers(html):
    phone_pattern = re.compile(r'(\+?[\d\s\-\(\)]{7,15})')
    raw = phone_pattern.findall(html)
    phones = []
    for p in raw:
        cleaned = p.strip()
        if len(cleaned) >= 7 and not is_duplicate(cleaned):
            phones.append(cleaned)
    return phones[:50]

def extract_tables(soup):
    tables = []
    for table in soup.find_all('table'):
        rows = []
        for row in table.find_all('tr'):
            cells = [cell.get_text().strip()
                     for cell in row.find_all(['td', 'th'])]
            if cells:
                rows.append(cells)
        if rows:
            tables.append(rows)
    return tables

def extract_paragraphs(soup):
    paragraphs = []
    for p in soup.find_all('p'):
        text = p.get_text().strip()
        if text and len(text) > 5 and not is_duplicate(text):
            paragraphs.append(text)
    return paragraphs

def extract_from_page(html, base_url, choice):
    soup = BeautifulSoup(html, 'html.parser')
    results = {}
    if choice in ["1", "8"]:
        results["headlines"] = extract_headlines(soup)
    if choice in ["2", "8"]:
        results["links"] = extract_links(soup, base_url)
    if choice in ["3", "8"]:
        results["images"] = extract_images(soup, base_url)
    if choice in ["4", "8"]:
        results["emails"] = extract_emails(html)
    if choice in ["5", "8"]:
        results["phones"] = extract_phone_numbers(html)
    if choice in ["6", "8"]:
        results["tables"] = extract_tables(soup)
    if choice in ["7", "8"]:
        results["paragraphs"] = extract_paragraphs(soup)
    return results

def merge_results(all_results):
    merged = {}
    for results in all_results:
        for key, value in results.items():
            if key not in merged:
                merged[key] = []
            merged[key].extend(value)
    return merged

# ─── Save Functions ───────────────────────────────────────

def save_json(data, filename):
    with open(filename + ".json", "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print("[+] JSON saved: " + filename + ".json")

def save_csv(data, filename):
    if not data:
        return
    with open(filename + ".csv", "w", newline="", encoding="utf-8") as f:
        if isinstance(data[0], dict):
            writer = csv.DictWriter(f, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        else:
            writer = csv.writer(f)
            for row in data:
                writer.writerow([row])
    print("[+] CSV saved: " + filename + ".csv")

def save_excel(results, filename):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sections = {
        "Headlines": results.get("headlines", []),
        "Links": results.get("links", []),
        "Images": results.get("images", []),
        "Emails": results.get("emails", []),
        "Phones": results.get("phones", []),
        "Paragraphs": results.get("paragraphs", []),
    }
    for sheet_name, data in sections.items():
        if not data:
            continue
        ws = wb.create_sheet(title=sheet_name)
        if isinstance(data[0], dict):
            headers = list(data[0].keys())
            ws.append(headers)
            for row in data:
                ws.append(list(row.values()))
        else:
            ws.append([sheet_name])
            for item in data:
                ws.append([item])
    if results.get("tables"):
        for i, table in enumerate(results["tables"]):
            ws = wb.create_sheet(title="Table_" + str(i + 1))
            for row in table:
                ws.append(row)
    excel_filename = filename + ".xlsx"
    wb.save(excel_filename)
    print("[+] Excel saved: " + excel_filename)

# ─── Main ─────────────────────────────────────────────────

def main():
    print("=" * 50)
    print("   Universal Web Scraper v2")
    print("   by ashardrach")
    print("=" * 50)

    url = input("\nEnter URL to scrape: ").strip()
    if not url.startswith("http"):
        url = "http://" + url

    parsed = urlparse(url)
    base_url = parsed.scheme + "://" + parsed.netloc

    print("\nScrape mode:")
    print("1. Single page only")
    print("2. Multi-page (follow pagination)")
    mode = input("Enter mode (1/2): ").strip()

    print("\nWhat do you want to extract?")
    print("1. Headlines")
    print("2. Links")
    print("3. Images")
    print("4. Emails")
    print("5. Phone numbers")
    print("6. Tables")
    print("7. Paragraphs")
    print("8. Everything")
    choice = input("\nEnter choice (1-8): ").strip()

    download = "n"
    if choice in ["3", "8"]:
        download = input("\nDownload images to disk? (y/n): ").strip().lower()

    max_pages = 10
    if mode == "2":
        try:
            max_pages = int(input("Max pages to scrape (default 10): ").strip())
        except:
            max_pages = 10

    reset_duplicates()

    if mode == "2":
        pages = get_all_pages(url, max_pages)
        all_results = []
        for page_url, html in pages:
            page_results = extract_from_page(html, base_url, choice)
            all_results.append(page_results)
        results = merge_results(all_results)
    else:
        html = fetch_page(url)
        if not html:
            return
        print("[+] Page fetched successfully.")
        results = extract_from_page(html, base_url, choice)

    if not results:
        print("[-] Nothing extracted.")
        return

    domain = parsed.netloc
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = "scrape_" + domain.replace(".", "_") + "_" + timestamp

    if download == "y" and results.get("images"):
        download_images(results["images"], domain)

    print("\n[*] Saving results...")
    save_json(results, filename)
    first_data = next((v for v in results.values() if v), [])
    save_csv(first_data, filename)
    save_excel(results, filename)

    print("\n" + "=" * 50)
    print("SCRAPE COMPLETE")
    print("=" * 50)
    for key, value in results.items():
        if isinstance(value, list):
            print("[+] " + key.capitalize() + ": " +
                  str(len(value)) + " found")
    print("=" * 50)

main()